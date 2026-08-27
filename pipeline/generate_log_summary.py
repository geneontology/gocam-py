#!/usr/bin/env python3
"""Generate Excel and HTML summaries of pipeline run results from JSONL logs."""

import itertools
import json
import logging
import re
from collections import defaultdict, namedtuple
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated, Any, Iterable

import pystow
import typer
import yaml
from _common import (
    PIPELINE_LOG_FORMAT_VERSION,
    PIPELINE_LOG_RECORD_TYPE,
    PIPELINE_STEP_ORDER,
    PipelineStep,
    normalize_model_id,
    setup_logger,
)
from jinja2 import Environment, FileSystemLoader, StrictUndefined
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.progress import Progress, track

from gocam import __version__

app = typer.Typer()

logger = logging.getLogger(__name__)

_PYSTOW_MODULE = pystow.module("gocam")
_CURRENT_GROUPS_YAML_URL = "https://current.geneontology.org/metadata/groups.yaml"
_CURRENT_USERS_YAML_URL = "https://current.geneontology.org/metadata/users.yaml"


@dataclass
class ProvidingGroup:
    """Resolved identity and display metadata for a model-providing group."""

    id: str
    label: str
    shorthand: str | None = None


@dataclass
class ModelSummary:
    """Merged reporting data for one model across all pipeline steps."""

    model_id: str
    title: str | None = None
    model_status: str | None = None
    date_modified: str | None = None
    contributors: set[str] = field(default_factory=set)
    contributor_names: list[str] = field(default_factory=list)
    provided_by: set[str] = field(default_factory=set)
    providing_groups: list[ProvidingGroup] = field(default_factory=list)
    groups: set[str] = field(default_factory=set)
    provenance_scope: str | None = None
    longest_path: int | None = None
    pipeline_status: str = "success"
    pipeline_status_details: str | None = None
    processed_steps: set[PipelineStep] = field(default_factory=set)
    warnings: list[Any] = field(default_factory=list)


@dataclass
class GroupReport:
    """Production GO-CAM-like models and pipeline results for one providing group."""

    group: ProvidingGroup
    models: list[ModelSummary]
    filename: str = ""

    @property
    def success_models(self) -> list[ModelSummary]:
        return [model for model in self.models if model.pipeline_status == "success"]

    @property
    def filtered_models(self) -> list[ModelSummary]:
        return [model for model in self.models if model.pipeline_status == "filtered"]

    @property
    def error_models(self) -> list[ModelSummary]:
        return [model for model in self.models if model.pipeline_status == "error"]


@dataclass(frozen=True)
class PipelineStepLog:
    """A pipeline step log file."""

    path: Path
    step: PipelineStep


def read_step_log_header(step_file: Path) -> PipelineStep:
    """Read and validate a pipeline step log header."""
    with step_file.open() as log_file:
        line = log_file.readline()

    if not line:
        raise ValueError(f"Pipeline log {step_file} is empty")
    try:
        header = json.loads(line)
    except json.JSONDecodeError as error:
        raise ValueError(f"Invalid JSON in {step_file} at line 1: {error}") from error

    if not isinstance(header, dict):
        raise ValueError(f"{step_file} pipeline log header must be a JSON object")
    if header.get("record_type") != PIPELINE_LOG_RECORD_TYPE:
        raise ValueError(f"{step_file} does not start with a pipeline log header")
    format_version = header.get("format_version")
    if type(format_version) is not int or format_version != PIPELINE_LOG_FORMAT_VERSION:
        raise ValueError(
            f"{step_file} uses unsupported pipeline log format version "
            f"{format_version!r}"
        )
    step_value = header.get("step")
    try:
        return PipelineStep(step_value)
    except ValueError as error:
        raise ValueError(
            f"{step_file} declares unknown pipeline step {step_value!r}"
        ) from error


def discover_step_logs(
    logs_dir: Path, extension: str = ".jsonl"
) -> list[PipelineStepLog]:
    """Find, validate, and canonically order pipeline step log files.

    Args:
        logs_dir: Path to the directory containing log files.
        extension: The file extension to filter log files (default: .jsonl).

    Returns:
        Logs ordered by their step's position in the canonical pipeline.
    """
    log_files = sorted(
        file
        for file in logs_dir.iterdir()
        if file.is_file() and file.suffix == extension and not file.name.startswith(".")
    )
    logs_by_step: dict[PipelineStep, PipelineStepLog] = {}
    for log_file in log_files:
        step = read_step_log_header(log_file)
        if step in logs_by_step:
            first_log = logs_by_step[step]
            raise ValueError(
                f"Found multiple logs for pipeline step {step.value!r}: "
                f"{first_log.path} and {log_file}"
            )
        logs_by_step[step] = PipelineStepLog(path=log_file, step=step)

    return [logs_by_step[step] for step in PIPELINE_STEP_ORDER if step in logs_by_step]


def iter_log_results(
    step_log: PipelineStepLog,
) -> Iterable[tuple[str, dict[str, Any]]]:
    """Iterate over the entries in a JSONL log file, yielding model_id and entry data.

    Args:
        step_log: Validated pipeline step log to read.

    Yields:
        Tuples of (normalized_model_id, entry_dict) for each entry in the log file
    """
    with step_log.path.open() as f:
        # Skip the header line
        next(f)
        for line_number, line in enumerate(f, start=2):
            try:
                entry = json.loads(line)
            except json.JSONDecodeError as error:
                raise ValueError(
                    f"Invalid JSON in {step_log.path} at line {line_number}: {error}"
                ) from error

            if not isinstance(entry, dict):
                raise ValueError(
                    f"Pipeline log result at line {line_number} must be a JSON "
                    f"object in {step_log.path}"
                )
            model_id = entry.get("model_id")
            if model_id is None:
                raise ValueError(
                    f"Missing model_id in {step_log.path} at line {line_number}"
                )
            if not isinstance(model_id, str):
                raise ValueError(
                    f"model_id in {step_log.path} at line {line_number} must be a "
                    "string"
                )
            yield normalize_model_id(model_id), entry


def hyperlink_formula(url: str, display_text: str) -> str:
    """Generate an Excel formula for a hyperlink.

    Args:
        url: The URL the hyperlink should point to.
        display_text: The text to display in the cell.

    Returns:
        A string containing the Excel formula for the hyperlink.
    """
    return f"""=HYPERLINK("{url}", "{display_text}")"""


def format_warning(warning: Any) -> str:
    """Format a warning dictionary into a string for display in Excel.

    Args:
        warning: A dictionary containing warning information, expected to have 'type' and 'message' keys.

    Returns:
        A formatted string representing the warning.
    """
    if isinstance(warning, str):
        return warning
    elif isinstance(warning, dict):
        warning_type = warning.get("type", "")
        message = warning.get("message", "")
        return f"{warning_type}{': ' if warning_type else ''}{message}"
    else:
        return str(warning)


def summarize_model_entries(
    model_id: str,
    entries: list[dict[str, Any]],
    *,
    processed_steps: set[PipelineStep],
) -> ModelSummary:
    """Merge report entries for one model into a renderer-independent summary."""
    summary = ModelSummary(model_id=model_id, processed_steps=set(processed_steps))
    for entry in entries:
        summary.warnings.extend(entry.get("warnings") or [])

        meta = entry.get("meta") or {}
        if summary.title is None:
            summary.title = meta.get("title")
        if summary.model_status is None:
            summary.model_status = meta.get("status")
        if summary.date_modified is None:
            summary.date_modified = meta.get("date_modified")
        if summary.provenance_scope is None:
            summary.provenance_scope = meta.get("provenance_scope")
        if summary.longest_path is None:
            summary.longest_path = meta.get("longest_path")
        summary.contributors.update(meta.get("contributors") or [])
        summary.provided_by.update(meta.get("provided_by") or [])
        summary.groups.update(meta.get("groups") or [])

        if summary.pipeline_status == "success" and entry.get("status") != "success":
            # This is the first non-success status we've seen, so record that status
            # for the report along with any associated reason and details.
            summary.pipeline_status = entry.get("status", "unknown")
            reason = entry.get("reason", "")
            details = entry.get("details", "")
            summary.pipeline_status_details = reason
            if details:
                if summary.pipeline_status_details:
                    summary.pipeline_status_details += f": {details}"
                else:
                    summary.pipeline_status_details = details

    return summary


def load_metadata_by_id(
    path: Path | None, *, default_url: str, id_field: str
) -> dict[str, dict[str, Any]]:
    """Load GO metadata from an explicit file or a cached default URL."""
    if path is None:
        path = _PYSTOW_MODULE.ensure(
            url=default_url,
            download_kwargs={"backend": "requests"},
        )

    with path.open() as metadata_file:
        records = yaml.safe_load(metadata_file)
    if not isinstance(records, list):
        raise ValueError(f"Expected a list of metadata records in {path}")

    return {
        str(record[id_field]): record
        for record in records
        if isinstance(record, dict) and record.get(id_field)
    }


def resolve_display_values(
    identifiers: set[str],
    metadata_by_id: dict[str, dict[str, Any]],
    *,
    display_field: str,
) -> list[str]:
    """Resolve identifiers to display values, retaining unknown identifiers."""
    return sorted(
        {
            str(metadata_by_id.get(identifier, {}).get(display_field) or identifier)
            for identifier in identifiers
        }
    )


def resolve_model_summary_metadata(
    summary: ModelSummary,
    *,
    users_by_id: dict[str, dict[str, Any]],
    groups_by_id: dict[str, dict[str, Any]],
) -> ModelSummary:
    """Add renderer-independent contributor and group display metadata."""
    summary.contributor_names = resolve_display_values(
        summary.contributors, users_by_id, display_field="nickname"
    )

    if summary.provided_by:
        summary.providing_groups = [
            ProvidingGroup(
                id=group_id,
                label=str(groups_by_id.get(group_id, {}).get("label") or group_id),
                shorthand=groups_by_id.get(group_id, {}).get("shorthand"),
            )
            for group_id in sorted(summary.provided_by)
        ]
    else:
        summary.providing_groups = [
            ProvidingGroup(id=group_label, label=group_label)
            for group_label in sorted(summary.groups)
        ]

    return summary


def slugify(value: str) -> str:
    """Convert a display value into a safe, readable filename component."""
    slug = re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")
    return slug or "group"


def sort_models_for_report(models: Iterable[ModelSummary]) -> list[ModelSummary]:
    """Sort models newest-first, using title and ID to order equal dates."""
    models_by_title = sorted(
        models, key=lambda model: ((model.title or "").casefold(), model.model_id)
    )
    return sorted(
        models_by_title,
        key=lambda model: model.date_modified or "",
        reverse=True,
    )


def build_group_reports(
    model_summaries: Iterable[ModelSummary],
) -> list[GroupReport]:
    """Group production GO-CAM-like model summaries by provider for HTML reporting.

    GO-CAM-like models are the ones that have been processed by the FILTER step (they
    were not filtered out by the "obvious not a GO-CAM" checks in the CONVERT step).
    """
    models_by_group: defaultdict[str, list[ModelSummary]] = defaultdict(list)
    groups_by_id: dict[str, ProvidingGroup] = {}
    unassigned = ProvidingGroup(
        id="urn:gocam-report:unassigned",
        label="Unassigned",
        shorthand="unassigned",
    )

    for summary in model_summaries:
        if (
            summary.model_status != "production"
            or PipelineStep.FILTER not in summary.processed_steps
        ):
            continue

        providing_groups = summary.providing_groups or [unassigned]
        for group in providing_groups:
            groups_by_id[group.id] = group
            models_by_group[group.id].append(summary)

    reports = [
        GroupReport(
            group=groups_by_id[group_id],
            models=sort_models_for_report(models),
        )
        for group_id, models in models_by_group.items()
    ]
    reports.sort(key=lambda report: (report.group.label.casefold(), report.group.id))

    for report in reports:
        filename_group = slugify(
            report.group.shorthand or report.group.label or report.group.id
        )
        report.filename = f"{filename_group}_go-cam_pipeline_report.html"

    return reports


def format_plural(count: int, singular: str, plural: str | None = None) -> str:
    """Return a string with the count and the appropriate singular/plural form."""
    if plural is None:
        plural = singular + "s"
    return f"{count} {singular if count == 1 else plural}"


_HTML_ENVIRONMENT = Environment(
    autoescape=True,
    loader=FileSystemLoader(
        Path(__file__).parent / "templates" / "generate_log_summary"
    ),
    undefined=StrictUndefined,
)


_HTML_ENVIRONMENT.filters["warning_text"] = format_warning
_HTML_ENVIRONMENT.filters["pluralize"] = format_plural


def render_excel_summary(
    model_summaries: list[ModelSummary],
    *,
    extra_metadata: list[str] | None,
    generated_at: str,
) -> Workbook:
    """Render all pipeline results into an Excel workbook."""
    Column = namedtuple("Column", ["name", "width", "definition"])
    columns = [
        Column(
            "Model ID",
            25,
            "The unique identifier for the model, without the `gomodel:` prefix.",
        ),
        Column(
            "VPE", 10, "Link to view the model in the Noctua Visual Pathway Editor."
        ),
        Column(
            "Graph Editor", 15, "Link to view the model in the Noctua Graph Editor."
        ),
        Column("Minerva JSON", 15, "Link to download the model's Minerva JSON file."),
        Column("Title", 55, "The title of the model"),
        Column("Model State", 13, "The state of the model"),
        Column("Date Modified", 15, "The date the model was last modified."),
        Column(
            "Pipeline Status",
            14,
            "The final status of the model after running through the pipeline. "
            "Values can be 'error' (meaning something unexpected happened while processing the "
            "model that caused a pipeline step to not complete), 'filtered' (meaning the model was "
            "processed by a pipeline step but did not meet our criteria to continue to the next "
            "pipeline step), or 'success' (meaning the model was processed successfully through all "
            "steps of the pipeline).",
        ),
        Column(
            "Pipeline Status Details",
            35,
            "Additional details about the pipeline status, such as error or filtering messages if "
            "the pipeline did not complete successfully.",
        ),
        Column(
            "Contributors",
            35,
            "All contributors found in the model's available provenance. Names are shown when "
            "available; otherwise contributor identifiers are shown.",
        ),
        Column(
            "Groups",
            35,
            "All groups that contributed to the model. Group labels are shown when available; "
            "otherwise the provider identifiers recorded in model provenance are shown.",
        ),
        Column(
            "Provenance Scope",
            18,
            "Whether contributor and group metadata covers all translated GO-CAM provenance or "
            "only top-level Minerva model provenance.",
        ),
        Column(
            "Longest Path",
            14,
            "The number of activities in the longest causal path through the model. Note that this "
            "information is computed late in the pipeline, so if a model was filtered out by an "
            "earlier step, this information may not be available.",
        ),
        Column(
            "Warning Count",
            15,
            "The total number of warnings generated for the model across all pipeline steps.",
        ),
        Column(
            "Warnings",
            200,
            "All warnings generated for the model across all pipeline steps, concatenated into "
            "a single cell with each warning on a new line.",
        ),
    ]

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Pipeline Summary"
    # Write header row
    summary_sheet.append([col.name for col in columns])

    # Set column widths
    for i, col in enumerate(columns, start=1):
        summary_sheet.column_dimensions[get_column_letter(i)].width = col.width

    # Style values for use later
    fill_green = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    alignment_wrapped = Alignment(wrap_text=True, vertical="top")
    font_bold = Font(bold=True)

    row = 1
    for summary in model_summaries:
        row += 1
        group_display_values = sorted(
            {group.label for group in summary.providing_groups}
        )
        formatted_contributors = (
            ", ".join(summary.contributor_names) if summary.contributor_names else None
        )
        formatted_groups = (
            ", ".join(group_display_values) if group_display_values else None
        )
        formatted_warnings = (
            "\n".join(format_warning(w) for w in summary.warnings)
            if summary.warnings
            else None
        )
        summary_sheet.append(
            [
                summary.model_id,
                hyperlink_formula(
                    f"http://noctua.geneontology.org/workbench/noctua-visual-pathway-editor/?model_id=gomodel:{summary.model_id}",
                    "VPE",
                ),
                hyperlink_formula(
                    f"http://noctua.geneontology.org/editor/graph/gomodel:{summary.model_id}",
                    "Graph Editor",
                ),
                hyperlink_formula(
                    f"https://go-public.s3.amazonaws.com/files/go-cam/{summary.model_id}.json",
                    "Minerva JSON",
                ),
                summary.title,
                summary.model_status,
                summary.date_modified,
                summary.pipeline_status,
                summary.pipeline_status_details,
                formatted_contributors,
                formatted_groups,
                summary.provenance_scope,
                summary.longest_path,
                len(summary.warnings),
                formatted_warnings,
            ]
        )
        if summary.pipeline_status == "success":
            for cell in summary_sheet[row]:
                cell.fill = fill_green

    # Create the Metadata worksheet
    metadata_sheet = wb.create_sheet("Metadata")
    metadata_sheet.column_dimensions["A"].width = 20
    metadata_sheet.column_dimensions["B"].width = 60

    # Add Provenance section with generation timestamp, software version, and any additional \
    # metadata provided via command-line arguments
    metadata_sheet.append(["Provenance"])
    for cell in metadata_sheet[metadata_sheet.max_row]:
        cell.font = font_bold
    metadata_sheet.append(["Generated on", generated_at])
    metadata_sheet.append(["gocam-py version", __version__])
    if extra_metadata:
        for item in extra_metadata:
            if "=" not in item:
                raise typer.BadParameter(
                    f"Invalid metadata entry {item!r}. Expected format: Key=Value."
                )
            key, value = item.split("=", 1)
            metadata_sheet.append([key.strip(), value.strip()])

    # Add column definitions section
    metadata_sheet.append([])
    metadata_sheet.append(["Column Definitions"])
    for cell in metadata_sheet[metadata_sheet.max_row]:
        cell.font = font_bold
    for col in columns:
        metadata_sheet.append([col.name, col.definition])

    for cell in metadata_sheet["B"]:
        cell.alignment = alignment_wrapped

    # Add filters to the header row
    last_column_letter = get_column_letter(len(columns))
    summary_sheet.auto_filter.ref = f"A1:{last_column_letter}{row}"

    # Apply text wrapping and alignment to all cells
    for row_cells in summary_sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = alignment_wrapped

    # Make the header row bold
    for cell in summary_sheet[1]:
        cell.font = font_bold

    return wb


def render_group_report(report: GroupReport, *, generated_at: str) -> str:
    """Render production GO-CAM-like model results for one providing group."""
    return _HTML_ENVIRONMENT.get_template("group.html.jinja2").render(
        report=report, generated_at=generated_at
    )


def render_group_index(reports: list[GroupReport], *, generated_at: str) -> str:
    """Render the index linking to every providing-group report."""
    return _HTML_ENVIRONMENT.get_template("index.html.jinja2").render(
        reports=reports, generated_at=generated_at
    )


@app.command()
def main(
    logs_dir: Annotated[
        Path,
        typer.Option(
            exists=True,
            file_okay=False,
            dir_okay=True,
            readable=True,
            help="Directory containing step JSONL log files.",
        ),
    ],
    log_file_extension: Annotated[
        str,
        typer.Option(
            help="File extension used to find log files in logs directory (default: .jsonl).",
        ),
    ] = ".jsonl",
    verbose: Annotated[
        int,
        typer.Option(
            "--verbose",
            "-v",
            count=True,
            help="Increase verbosity level. Can be used multiple times.",
        ),
    ] = 0,
    limit: Annotated[
        int,
        typer.Option(
            help="Limit the number of models included in the generated summary."
        ),
    ] = 0,
    metadata: Annotated[
        list[str] | None,
        typer.Option(
            help="Additional info to include in the metadata sheet, in 'Key=Value' format. Can be used multiple times.",
        ),
    ] = None,
    excel_output: Annotated[
        Path | None,
        typer.Option(
            file_okay=True,
            dir_okay=False,
            writable=True,
            help="File to write the generated Excel summary to.",
        ),
    ] = None,
    html_output_dir: Annotated[
        Path | None,
        typer.Option(
            file_okay=False,
            dir_okay=True,
            help="Directory to write static HTML reports to.",
        ),
    ] = None,
    goc_users_yaml: Annotated[
        Path | None,
        typer.Option(
            exists=True,
            file_okay=True,
            dir_okay=False,
            readable=True,
            help="YAML file defining GOC users. Defaults to current.geneontology.org metadata.",
        ),
    ] = None,
    goc_groups_yaml: Annotated[
        Path | None,
        typer.Option(
            exists=True,
            file_okay=True,
            dir_okay=False,
            readable=True,
            help="YAML file defining GOC groups. Defaults to current.geneontology.org metadata.",
        ),
    ] = None,
) -> None:
    setup_logger(verbose)

    if excel_output is None and html_output_dir is None:
        raise typer.BadParameter(
            "At least one of --excel-output or --html-output-dir must be provided."
        )

    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    if excel_output and not excel_output.name.endswith(".xlsx"):
        raise typer.BadParameter("Output file must have .xlsx extension.")

    step_logs = discover_step_logs(logs_dir, log_file_extension)
    if not step_logs:
        raise typer.BadParameter(
            f"No log files with extension {log_file_extension} found in directory: {logs_dir}"
        )

    step_results_by_model_id: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    processed_steps_by_model_id: defaultdict[str, set[PipelineStep]] = defaultdict(set)
    for step_log in step_logs:
        for model_id, entry in iter_log_results(step_log):
            step_results_by_model_id[model_id].append(entry)
            processed_steps_by_model_id[model_id].add(step_log.step)

    users_by_id = load_metadata_by_id(
        goc_users_yaml, default_url=_CURRENT_USERS_YAML_URL, id_field="uri"
    )
    groups_by_id = load_metadata_by_id(
        goc_groups_yaml, default_url=_CURRENT_GROUPS_YAML_URL, id_field="id"
    )

    model_entries = step_results_by_model_id.items()
    if limit > 0:
        model_entries = itertools.islice(model_entries, limit)

    model_summaries = [
        resolve_model_summary_metadata(
            summarize_model_entries(
                model_id,
                entries,
                processed_steps=processed_steps_by_model_id[model_id],
            ),
            users_by_id=users_by_id,
            groups_by_id=groups_by_id,
        )
        for model_id, entries in track(
            model_entries, description="Building log summaries..."
        )
    ]

    if excel_output is not None:
        with Progress() as progress:
            task = progress.add_task(description="Rendering Excel summary...", total=1)
            wb = render_excel_summary(
                model_summaries, extra_metadata=metadata, generated_at=generated_at
            )
            wb.save(excel_output)
            progress.advance(task, 1)

    if html_output_dir is not None:
        reports = build_group_reports(model_summaries)
        html_output_dir.mkdir(parents=True, exist_ok=True)
        with Progress() as progress:
            task = progress.add_task(
                description="Writing HTML reports...", total=len(reports) + 1
            )
            (html_output_dir / "index.html").write_text(
                render_group_index(reports, generated_at=generated_at), encoding="utf-8"
            )
            progress.advance(task, 1)
            for report in reports:
                (html_output_dir / report.filename).write_text(
                    render_group_report(report, generated_at=generated_at),
                    encoding="utf-8",
                )
                progress.advance(task, 1)


if __name__ == "__main__":
    app()
